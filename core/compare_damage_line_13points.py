from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = PROJECT_ROOT / "output"
DISPATCH_LINES_CSV = OUTPUT_DIR / "dispatch_chart_lines.csv"
DEAD_WATER_CSV = OUTPUT_DIR / "dead_water_level_results.csv"
STORAGE_CURVE_CSV = PROJECT_ROOT / "data" / "water_level_storage.csv"
TEACHER_DIR = Path("/Users/lupeng/Downloads")

MONTH_LABELS = ["3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "1月", "2月", "3月"]
OUR_12_MONTHS = ["4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "1月", "2月", "3月"]


def read_dispatch_rows() -> dict[float, dict[str, dict[str, str]]]:
    by_level: dict[float, dict[str, dict[str, str]]] = {}
    with DISPATCH_LINES_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            by_level.setdefault(float(row["正常蓄水位_m"]), {})[row["月份"]] = row
    return by_level


def read_dead_storage() -> dict[float, float]:
    with STORAGE_CURVE_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        curve_rows = list(csv.DictReader(f))
    level_points = [float(row["水位_m"]) for row in curve_rows]
    storage_points = [float(row["库容_m3每秒月"]) for row in curve_rows]

    def interpolate(x: float) -> float:
        if x <= level_points[0]:
            return storage_points[0]
        if x >= level_points[-1]:
            return storage_points[-1]
        for i in range(1, len(level_points)):
            if x <= level_points[i]:
                x0, x1 = level_points[i - 1], level_points[i]
                y0, y1 = storage_points[i - 1], storage_points[i]
                return y0 + (x - x0) * (y1 - y0) / (x1 - x0)
        return storage_points[-1]

    dead_storage: dict[float, float] = {}
    with DEAD_WATER_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            normal_level = float(row["正常蓄水位_m"])
            dead_storage[normal_level] = interpolate(float(row["死水位_m"]))
    return dead_storage


def read_teacher_storage(level: float) -> list[float]:
    workbook = load_workbook(TEACHER_DIR / f"Out_防破坏线_{level:.0f}.xlsx", data_only=True)
    sheet = workbook.active
    values: list[float] = []
    for month, storage in sheet.iter_rows(min_row=2, values_only=True):
        if month is None or storage is None:
            continue
        values.append(float(storage))
    if len(values) != 13:
        raise RuntimeError(f"老师 {level:.0f}m 防破坏线不是13个点: {len(values)}")
    return values


def write_comparison_csv(rows: list[dict[str, float | int | str]]) -> Path:
    output_path = OUTPUT_DIR / "dispatch_charts" / "teacher_damage_line_comparison_13points.csv"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "正常蓄水位_m",
                "序号",
                "横轴标签",
                "老师库蓄水量_m3s月",
                "我们库蓄水量_m3s月",
                "差值_我们减老师",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({key: round(value, 3) if isinstance(value, float) else value for key, value in row.items()})
    return output_path


def svg_polyline(points: list[tuple[float, float]], color: str, width: float = 2.0, dash: bool = False) -> str:
    dash_text = ' stroke-dasharray="6 4"' if dash else ""
    point_text = " ".join(f"{x:.1f},{y:.1f}" for x, y in points)
    return f'<polyline points="{point_text}" fill="none" stroke="{color}" stroke-width="{width}"{dash_text}/>'


def plot_level(level: float, rows: list[dict[str, float | int | str]]) -> Path:
    output_dir = OUTPUT_DIR / "dispatch_charts" / "teacher_comparison_charts"
    output_dir.mkdir(parents=True, exist_ok=True)
    teacher = [float(row["老师库蓄水量_m3s月"]) for row in rows]
    ours = [float(row["我们库蓄水量_m3s月"]) for row in rows]
    all_values = teacher + ours

    width, height = 1080, 620
    left, right, top, bottom = 80, 35, 52, 90
    plot_w, plot_h = width - left - right, height - top - bottom
    y_min, y_max = min(all_values) - 80, max(all_values) + 80

    def px(index: int) -> float:
        return left + plot_w * index / (len(MONTH_LABELS) - 1)

    def py(value: float) -> float:
        return top + (y_max - value) / (y_max - y_min) * plot_h

    diffs = [ours_value - teacher_value for ours_value, teacher_value in zip(ours, teacher)]
    elements = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="white"/>',
        f'<text x="{width / 2:.0f}" y="30" text-anchor="middle" font-size="22">正常蓄水位 {level:.0f}m 防破坏线对比（13点口径）</text>',
        f'<line x1="{left}" y1="{top}" x2="{left}" y2="{height-bottom}" stroke="#444"/>',
        f'<line x1="{left}" y1="{height-bottom}" x2="{width-right}" y2="{height-bottom}" stroke="#444"/>',
    ]
    for tick in range(6):
        value = y_min + (y_max - y_min) * tick / 5
        y = py(value)
        elements.append(f'<line x1="{left}" y1="{y:.1f}" x2="{width-right}" y2="{y:.1f}" stroke="#e3e3e3"/>')
        elements.append(f'<text x="{left-8}" y="{y+4:.1f}" text-anchor="end" font-size="12">{value:.0f}</text>')
    for index, label in enumerate(MONTH_LABELS):
        x = px(index)
        elements.append(f'<text x="{x:.1f}" y="{height-bottom+26}" text-anchor="middle" font-size="13">{label}</text>')

    teacher_points = [(px(index), py(value)) for index, value in enumerate(teacher)]
    our_points = [(px(index), py(value)) for index, value in enumerate(ours)]
    elements.append(svg_polyline(our_points, "#1f77b4", 2.4))
    elements.append(svg_polyline(teacher_points, "#d62728", 2.2, True))
    for x, y in our_points:
        elements.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4" fill="#1f77b4"/>')
    for x, y in teacher_points:
        elements.append(f'<rect x="{x-3.5:.1f}" y="{y-3.5:.1f}" width="7" height="7" fill="#d62728"/>')

    legend_x, legend_y = width - 205, 74
    elements.append(svg_polyline([(legend_x, legend_y), (legend_x + 34, legend_y)], "#1f77b4", 2.4))
    elements.append(f'<text x="{legend_x+43}" y="{legend_y+4}" font-size="13">我们的结果</text>')
    elements.append(svg_polyline([(legend_x, legend_y + 24), (legend_x + 34, legend_y + 24)], "#d62728", 2.2, True))
    elements.append(f'<text x="{legend_x+43}" y="{legend_y+28}" font-size="13">老师成果</text>')
    elements.append(
        f'<text x="{left+8}" y="{top+20}" font-size="13">平均偏差 {sum(abs(value) for value in diffs) / len(diffs):.1f}，最大偏差 {max(abs(value) for value in diffs):.1f}</text>'
    )
    elements.append(f'<text x="24" y="{height/2:.0f}" transform="rotate(-90 24,{height/2:.0f})" text-anchor="middle" font-size="14">库蓄水量 (m3/s·月)</text>')
    elements.append("</svg>")

    output_path = output_dir / f"teacher_vs_ours_damage_line_{level:.0f}m_13points.svg"
    output_path.write_text("\n".join(elements), encoding="utf-8")
    return output_path


def main() -> None:
    dispatch_rows = read_dispatch_rows()
    dead_storage = read_dead_storage()
    comparison_rows: list[dict[str, float | int | str]] = []
    plot_paths: list[Path] = []

    for level in [120.0, 115.0, 108.0, 100.0]:
        our_values = [float(dispatch_rows[level][month]["防破坏线库蓄水量_m3s月"]) for month in OUR_12_MONTHS]
        our_values.append(dead_storage[level])
        teacher_values = read_teacher_storage(level)
        level_rows: list[dict[str, float | int | str]] = []
        for index, (label, teacher_value, our_value) in enumerate(zip(MONTH_LABELS, teacher_values, our_values)):
            row = {
                "正常蓄水位_m": level,
                "序号": index,
                "横轴标签": label,
                "老师库蓄水量_m3s月": teacher_value,
                "我们库蓄水量_m3s月": our_value,
                "差值_我们减老师": our_value - teacher_value,
            }
            comparison_rows.append(row)
            level_rows.append(row)
        plot_paths.append(plot_level(level, level_rows))

    csv_path = write_comparison_csv(comparison_rows)
    print(f"13点对比表已写入: {csv_path}")
    for path in plot_paths:
        print(f"13点对比图已写入: {path}")


if __name__ == "__main__":
    main()
